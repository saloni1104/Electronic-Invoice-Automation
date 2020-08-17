import pandas as pd
import json
import xlsxwriter
import requests
import openpyxl as xl

#FORM JSON

workbook = xlsxwriter.Workbook('Sheet2.xlsx')
worksheet = workbook.add_worksheet()

res=requests.get("https://s1v6m61msh.execute-api.us-east-1.amazonaws.com/demo/retrievedocumentanalysisresult?Bucket=746479098454-scanned-documents&Document=Sample6.pdf&ResultType=FORM")

match_histories = res.json()

res1 = list(match_histories.keys())[12]
list1 = []
for i in range(len(match_histories[res1])):
    res2 = list(match_histories[res1].keys())[i]
    a = match_histories[res1][res2]
    for j in range(len(a)):
        if len(res2)!=0:
            res3 = res2.lower()
        a = match_histories[res1][res2]
        if 'state' in res3:
            if 'state' not in list1:
                worksheet.write(0,0, a[j])
                list1.append('state')
            else:
                continue
        if 'gstin' in res3:
            if 'gstin' not in list1:
                worksheet.write(0,1, a[j])
                list1.append('gstin')
            else:
                continue
        if 'invoice no' in res3:
            if res3 not in list1:
                worksheet.write(0,2, a[j])
                list1.append(res3)
            else:
                continue
        if 'date' in res3:
            if 'date' not in list1:
                worksheet.write(0,3, a[j])
                list1.append('date')
            else:
                continue
        if 'address' in res3:
            if 'address' not in list1:
                worksheet.write(0,4, a[j])
                list1.append('address')
            else:
                continue
        if 'po no' in res3 or 'po number' in res3 or 'p.o. number' in res3 or 'p.o. no' in res3:
            if 'po no' not in list1 or 'po number' not in list1:
                worksheet.write(0,5, a[j])
                list1.append('po no')
            else:
                continue
        if 'exp' in res3:
            if 'exp' not in list1:
                worksheet.write(0,6, a[j])
                list1.append('exp')
            else:
                continue
workbook.close()

#FORM JSON IN EXCEL TO ACTUAL INVOICE

filename ="C:\\Users\\salon\\OneDrive\\Desktop\\Sheet2.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0]

filename1 ="C:\\Users\\salon\\OneDrive\\Desktop\\Invoice_template_output_case_study(1).xlsx"
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.active

mc = ws1.max_column

c1 = ws1.cell(row = 1, column = 1).value
ws2.cell(row = 2, column = 2).value = c1
c2 = ws1.cell(row = 1, column = 2).value
ws2.cell(row = 6, column = 2).value = c2
c3 = ws1.cell(row = 1, column = 3).value
ws2.cell(row = 2, column = 4).value = c3
c4 = ws1.cell(row = 1, column = 4).value
ws2.cell(row = 3, column = 4).value = c4
c5 = ws1.cell(row = 1, column = 5).value
ws2.cell(row = 5, column = 2).value = c5
c6 = ws1.cell(row = 1, column = 6).value
ws2.cell(row = 9, column = 4).value = c6
c7 = ws1.cell(row = 1, column = 7).value
ws2.cell(row = 4, column = 4).value = c7

wb2.save(str(filename1))

#TABLE JSON TO EXCEL

j = 0
workbook = xlsxwriter.Workbook('Sheet1.xlsx')
worksheet = workbook.add_worksheet()

res=requests.get("https://s1v6m61msh.execute-api.us-east-1.amazonaws.com/demo/retrievedocumentanalysisresult?Bucket=746479098454-scanned-documents&Document=Sample6.pdf&ResultType=TABLE")

match_histories = res.json()

res1 = list(match_histories.keys())[12]

data1 = json.dumps(match_histories[res1], indent = 4, sort_keys=True)
final_dictionary = json.loads(data1)

one = final_dictionary[0]

res2 = list(one.keys())[0]

data2 = json.dumps(one[res2]["tr"], indent = 4, sort_keys=True)
final_dictionary1 = json.loads(data2)

for i in range(len(final_dictionary1)):
    k = 0
    two = final_dictionary1[i]
    for n in two["td"]:
        worksheet.write(j,k,n)
        k += 1
    j += 1
workbook.close()

#TABLE JSON IN EXCEL TO ACTUAL INVOICE

filename ="C:\\Users\\salon\\OneDrive\\Desktop\\Sheet1.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0]

filename1 ="C:\\Users\\salon\\OneDrive\\Desktop\\Invoice_template_output_case_study(1).xlsx"
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.active

mr = ws1.max_row 
mc = ws1.max_column
#print(mr, mc)

for i in range (1, mc + 1):
    value = ws1.cell(row = 1, column = i).value
    k = 15
    if len(value)!=0:
        value1 = value.lower()
    if "id" in value1 or "product id" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            if c.isnumeric():
                ws2.cell(row = k, column = 2).value = c
                k +=1
    elif "hsn" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 4).value = c
            k +=1
    elif "deseription" in value1 or "material" in value1 or "description" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 5).value = c
            k +=1
    elif "qty" in value1 or "quantity" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 6).value = c
            k +=1
    elif "unit" in value1 or "mrp" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 7).value = c
            k +=1
    elif "total" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 15).value = c
            k +=1
    elif "discount" in value1 or "dist" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 9).value = c
            k +=1
    elif "app" in value1 or "avg" in value1 or "average" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 16).value = c
            k +=1
    elif "igst" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 12).value = c
            k +=1
    elif "tax" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 8).value = c
            k +=1
    elif "cgst" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 11).value = c
            k +=1
    elif "sgst" in value1 or "utgst" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 10).value = c
            k +=1
    elif "cess" in value1:
        for j in range (2, mr + 1):
            c = ws1.cell(row = j, column = i).value
            ws2.cell(row = k, column = 13).value = c
            k +=1
wb2.save(str(filename1))
